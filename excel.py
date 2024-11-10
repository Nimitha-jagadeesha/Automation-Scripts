import pandas as pd
import openpyxl as pxl
from openpyxl.utils.dataframe import dataframe_to_rows



# Name of the workbook we'll be using
filename = 'Data.xlsx'

# Load the workbook (i.e. the existing workbook in your case)
excel_book = pxl.load_workbook(filename)


ws = excel_book['Data']
n = ws.max_row
print(n)
FirstName = []
LastName=[]
Marks1=[]
Marks2=[]
Formula1=[]
Formula2=[]
for i in range(n-1):
    FirstName.append(ws.cell(row=i+2,column = 1).value)
    LastName.append(ws.cell(row=i+2,column = 2).value)
    Marks1.append(ws.cell(row=i+2,column = 3).value)
    Marks2.append(ws.cell(row=i+2,column=4).value)
    Formula1.append('=A'+str(i+2)+'&" "&'+'B'+str(i+2))
    Formula2.append('=SUM(D'+str(i+2)+',E'+str(i+2)+')')


excelData = pd.DataFrame({
    'First Name': FirstName,
    'Last Name':LastName,
    'Full Name':Formula1,
    "Marks1":Marks1,
    "Marks2":Marks2,
    "Total":Formula2
})

rows = dataframe_to_rows(excelData, index=False)

# Create a new worksheet
excel_book.create_sheet('calculated')
# Work with the new worksheet
ws = excel_book['calculated']

# For each row
for r_idx, row in enumerate(rows, 1):
    # Write each cell for each column
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# Overwrite the workbook, now with two worksheets populated
excel_book.save(filename)
